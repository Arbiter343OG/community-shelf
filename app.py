from flask import Flask, render_template, request, redirect, jsonify, url_for, flash, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
import os
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

# Load env
load_dotenv()

app = Flask(__name__)

app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'dev-secret-key-123')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///shelf.db')
app.config['VOLUNTEER_ACCESS_CODE'] = os.getenv('VOLUNTEER_ACCESS_CODE', 'JOIN-SHELF-2026')

db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'welcome'

# MODELS
class Organization(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    org_access_code = db.Column(db.String(50), nullable=False)
    description = db.Column(db.String(255), default="Supporting our local community.")

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    org_id = db.Column(db.Integer, db.ForeignKey('organization.id'), nullable=True)
    organization = db.relationship('Organization', backref='members')

class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(50), default="General 📦")
    quantity = db.Column(db.Integer, default=1)
    unit = db.Column(db.String(20), default="units")
    low_threshold = db.Column(db.Integer, default=5)
    date_added = db.Column(db.DateTime, default=datetime.utcnow)
    org_id = db.Column(db.Integer, db.ForeignKey('organization.id'), nullable=False)

class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    username = db.Column(db.String(50))
    item_name = db.Column(db.String(100))
    action = db.Column(db.String(50))
    change = db.Column(db.String(20))
    org_id = db.Column(db.Integer, db.ForeignKey('organization.id'), nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        db.session.add(User(username='admin', password=generate_password_hash('123', method='scrypt'), role='volunteer'))
        db.session.add(User(username='guest', password=generate_password_hash('123', method='scrypt'), role='donor'))
        db.session.commit()

# ROUTES

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username')).first()
        if user and check_password_hash(user.password, request.form.get('password')):
            login_user(user)
            return redirect(url_for('index'))
        flash('Login Failed. Check your credentials.')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        reg_type = request.form.get('reg_type')

        if User.query.filter_by(username=username).first():
            flash('Username taken.')
            return redirect(url_for('register'))

        if reg_type == 'create':
            org_name = request.form.get('org_name')
            new_access_code = request.form.get('new_access_code')
            new_org = Organization(name=org_name, org_access_code=new_access_code)
            db.session.add(new_org)
            db.session.commit()
            target_org_id = new_org.id
            user_role = 'founder'
        else:
            access_code = request.form.get('access_code')
            org = Organization.query.filter_by(org_access_code=access_code).first()
            if not org:
                flash('Invalid Organization Access Code.')
                return redirect(url_for('register'))
            target_org_id = org.id
            user_role = role

        hashed_pw = generate_password_hash(password, method='scrypt')
        new_user = User(username=username, password=hashed_pw, role=user_role, org_id=target_org_id)
        db.session.add(new_user)
        db.session.commit()

        flash(f'Success! Welcome to {Organization.query.get(target_org_id).name}')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('welcome'))

@app.route('/')
@login_required
def index():
    search_query = request.args.get('q', '')
    base_query = Item.query.filter_by(org_id=current_user.org_id)

    if search_query:
        items = base_query.filter(
            (Item.name.contains(search_query)) |
            (Item.category.contains(search_query))
        ).all()
    else:
        items = base_query.order_by(Item.date_added.desc()).all()

    logs = ActivityLog.query.filter_by(org_id=current_user.org_id)\
        .order_by(ActivityLog.timestamp.desc()).limit(10).all()

    return render_template('index.html', items=items, search_query=search_query, user=current_user, logs=logs)

@app.route('/welcome')
def welcome():
    return render_template('welcome.html', user=current_user)

@app.route('/add', methods=['POST'])
@login_required
def add_item():
    if not current_user.org_id:
        flash("User is not associated with an organization.")
        return redirect(url_for('index'))

    name = request.form.get('name')
    category = request.form.get('category')
    unit = request.form.get('unit')
    threshold = int(request.form.get('threshold', 5))

    new_item = Item(
        name=name,
        category=category,
        unit=unit,
        org_id=current_user.org_id,
        low_threshold=threshold
    )
    db.session.add(new_item)

    db.session.add(ActivityLog(
        username=current_user.username,
        item_name=name,
        action="Added Item",
        change="+1",
        org_id=current_user.org_id
    ))

    db.session.commit()
    return redirect(url_for('index'))

@app.route('/update/<int:id>/<string:action>', methods=['POST'])
@login_required
def update(id, action):
    item = Item.query.get_or_404(id)

    # CRITICAL FIX: org check
    if item.org_id != current_user.org_id:
        return jsonify({'error': 'Unauthorized'}), 403

    if current_user.role == 'donor' and action == 'decrease':
        return jsonify({'error': 'Donors cannot remove stock'}), 403

    if action == 'increase':
        item.quantity += 1
    elif action == 'decrease' and item.quantity > 0:
        item.quantity -= 1

    db.session.add(ActivityLog(
        username=current_user.username,
        item_name=item.name,
        action="Updated Qty",
        change="+1" if action == 'increase' else "-1",
        org_id=current_user.org_id
    ))

    db.session.commit()
    return jsonify({'new_qty': item.quantity})

@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_item(id):
    item = Item.query.get_or_404(id)

    if item.org_id != current_user.org_id:
        return jsonify({'error': 'Unauthorized'}), 403

    if current_user.role not in ['volunteer', 'founder']:
        return jsonify({'error': 'Unauthorized'}), 403

    db.session.delete(item)

    db.session.add(ActivityLog(
        username=current_user.username,
        item_name=item.name,
        action="Deleted Item",
        change="-1",
        org_id=current_user.org_id
    ))

    db.session.commit()
    return jsonify({'success': True})

@app.route('/edit/<int:id>', methods=['POST'])
@login_required
def edit_item(id):
    item = Item.query.get_or_404(id)

    if item.org_id != current_user.org_id:
        return redirect(url_for('index'))

    if current_user.role not in ['volunteer', 'founder']:
        return redirect(url_for('index'))

    item.name = request.form.get('name')
    item.category = request.form.get('category')
    item.low_threshold = int(request.form.get('threshold', 5))

    db.session.commit()
    return redirect(url_for('index'))

@app.route('/export_audit_xlsx')
@login_required
def export_audit_xlsx():
    if current_user.role not in ['volunteer', 'founder']:
        return "Unauthorized", 403

    logs = ActivityLog.query.filter_by(org_id=current_user.org_id)\
        .order_by(ActivityLog.timestamp.desc()).all()

    items = Item.query.filter_by(org_id=current_user.org_id).all()

    # Map item name → unit
    item_units = {item.name: item.unit for item in items}

    wb = Workbook()

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # =========================
    # 📄 SHEET 1: AUDIT TRAIL
    # =========================
    ws = wb.active
    ws.title = "Audit Trail"

    headers = ['Timestamp', 'User', 'Action', 'Item', 'Unit', 'Change']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for log in logs:
        unit = item_units.get(log.item_name, "")
        ws.append([
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.username,
            log.action.upper(),
            log.item_name,
            unit,
            log.change
        ])

    # Auto width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # =========================
    # 📊 DATA PROCESSING
    # =========================
    total_added = 0
    total_removed = 0

    user_activity = {}
    item_activity_count = {}
    item_quantity_flow = {}

    for log in logs:
        # Count adds/removes
        if log.change == "+1":
            total_added += 1
        elif log.change == "-1":
            total_removed += 1

        # User activity
        user_activity[log.username] = user_activity.get(log.username, 0) + 1

        # Item key with unit
        unit = item_units.get(log.item_name, "")
        item_key = f"{log.item_name} ({unit})" if unit else log.item_name

        # Count-based activity
        item_activity_count[item_key] = item_activity_count.get(item_key, 0) + 1

        # Quantity-based flow
        if item_key not in item_quantity_flow:
            item_quantity_flow[item_key] = 0

        if log.change == "+1":
            item_quantity_flow[item_key] += 1
        elif log.change == "-1":
            item_quantity_flow[item_key] -= 1

    net_change = total_added - total_removed

    most_active_user = max(user_activity, key=user_activity.get) if user_activity else "N/A"
    most_active_item = max(item_activity_count, key=item_activity_count.get) if item_activity_count else "N/A"

    # =========================
    # 📊 SHEET 2: SUMMARY
    # =========================
    summary_ws = wb.create_sheet(title="Summary")

    summary_data = [
        ["Metric", "Value"],
        ["Total Items Added", total_added],
        ["Total Items Removed", total_removed],
        ["Net Stock Change", net_change],
        ["Most Active User", most_active_user],
        ["Most Moved Item", most_active_item],
    ]

    for row in summary_data:
        summary_ws.append(row)

    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # =========================
    # 📈 SHEET 3: INSIGHTS
    # =========================
    insights_ws = wb.create_sheet(title="Insights")

    # Top Users
    insights_ws.append(["Top Users", "Activity Count"])
    sorted_users = sorted(user_activity.items(), key=lambda x: x[1], reverse=True)[:5]

    for user, count in sorted_users:
        insights_ws.append([user, count])

    insights_ws.append([])

    # Top Items (by activity count)
    insights_ws.append(["Top Items (by Actions)", "Count"])
    sorted_items = sorted(item_activity_count.items(), key=lambda x: x[1], reverse=True)[:5]

    for item, count in sorted_items:
        insights_ws.append([item, count])

    insights_ws.append([])

    # Top Items (by quantity flow)
    insights_ws.append(["Top Items (Net Quantity Movement)", "Net Change"])
    sorted_flow = sorted(item_quantity_flow.items(), key=lambda x: abs(x[1]), reverse=True)[:5]

    for item, qty in sorted_flow:
        insights_ws.append([item, qty])

    for cell in insights_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # =========================
    # 📁 EXPORT
    # =========================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    org_name = current_user.organization.name.replace(' ', '_') if current_user.organization else "Admin"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Shelf_Report_{org_name}_{timestamp}.xlsx"
    )

@app.route('/team')
@login_required
def team_management():
    if current_user.role not in ['volunteer', 'founder']:
        flash("Access Denied.")
        return redirect(url_for('index'))

    all_members = User.query.filter_by(org_id=current_user.org_id).all()
    volunteers = [u for u in all_members if u.role in ['volunteer', 'founder']]
    donors = [u for u in all_members if u.role == 'donor']

    logs = ActivityLog.query.filter_by(org_id=current_user.org_id)\
        .order_by(ActivityLog.timestamp.desc()).limit(10).all()

    return render_template('team.html', volunteers=volunteers, donors=donors, logs=logs)

@app.route('/remove_member/<int:user_id>', methods=['POST'])
@login_required
def remove_member(user_id):
    if current_user.role.lower() != 'founder':
        return jsonify({'error': 'Only the Founder can remove members'}), 403

    user_to_remove = User.query.get_or_404(user_id)

    if user_to_remove.org_id == current_user.org_id and user_to_remove.id != current_user.id:
        db.session.delete(user_to_remove)
        db.session.commit()
        return jsonify({'success': True})

    return jsonify({'error': 'Invalid request'}), 400

@app.route('/update_mission', methods=['POST'])
@login_required
def update_mission():
    if current_user.role.lower() != 'founder':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    current_user.organization.description = data.get('description')
    db.session.commit()

    return jsonify({'success': True})

if __name__ == "__main__":
    app.run(debug=True)
